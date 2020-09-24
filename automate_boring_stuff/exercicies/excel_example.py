import openpyxl

wb = openpyxl.load_workbook("example.xlsx")

print(type(wb))

print(wb.sheetnames) # The workbook's sheets' names.

sheet = wb['Planilha1'] # Get a sheet from the workbook.

sheet.title # Get the sheet's title as a string.

anotherSheet = wb.active # Get the active sheet.

sheet['A1'] # Get a cell from the sheet.

sheet['A1'].value # Get the value from the cell.

c = sheet["B1"]

# Get the row, column, and value from the cell.
'Row %s, Column %s is %s' % (c.row, c.column, c.value)

'Cell %s is %s' % (c.coordinate, c.value) # 'Cell B1 is Apples'

# Using Integers to acess cells
sheet.cell(row=1, column=2)
sheet.cell(row=1, column=2).value

print(sheet.max_row) # Get the highest row number.

print(sheet.max_column) # Get the highest column number.

from openpyxl.utils import get_column_letter, column_index_from_string

get_column_letter(1) # Translate column 1 to a letter.

column_index_from_string('A') # Get A's number.

tuple(sheet['A1':'C3']) # Get all cells from A1 to C3.

# Creating and Removing Sheets

wb = openpyxl.Workbook()

wb.create_sheet() # Add a new sheet.

# Create a new sheet at index 0.
wb.create_sheet(index=0, title='First Sheet')

del wb['First Sheet']

wb.save('example_copy.xlsx')